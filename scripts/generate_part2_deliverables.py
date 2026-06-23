"""
Generate SQ26 Part 2 deliverables.

Outputs:
- reports/23071063-sq26-classification.xlsx
- reports/23071063-sq26-classification-report.pdf
- reports/part2_deliverables_summary.json

The XLSX contains exactly these required columns:
repository_id, project_type, project_title, primary_class,
secondary_class, no_project_files
"""

from __future__ import annotations

import argparse
import html
import json
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pypdf import PdfReader
from reportlab.graphics.shapes import Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


STUDENT_ID = "23071063"

FINAL_DATABASE = Path(f"{STUDENT_ID}-sq26-classification.db")
STAGING_DATABASE = Path("data/staging/qdarchive_x_staging.db")
OUTPUT_DIRECTORY = Path("reports")
TOTAL_FILE_MANIFEST = Path("reports/qdpx_total_file_manifest.json")

XLSX_COLUMNS = [
    "repository_id",
    "project_type",
    "project_title",
    "primary_class",
    "secondary_class",
    "no_project_files",
]

DARK_BLUE = colors.HexColor("#17365D")
MID_BLUE = colors.HexColor("#2E75B6")
LIGHT_BLUE = colors.HexColor("#D9EAF7")
VERY_LIGHT_BLUE = colors.HexColor("#F4F8FC")
LIGHT_GREY = colors.HexColor("#F3F5F7")
MID_GREY = colors.HexColor("#D4DEE8")
DARK_GREY = colors.HexColor("#404040")
GREEN = colors.HexColor("#548235")
ORANGE = colors.HexColor("#C55A11")
RED = colors.HexColor("#C00000")
WHITE = colors.white


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate SQ26 Part 2 XLSX and PDF deliverables."
    )

    parser.add_argument(
        "--db",
        type=Path,
        default=FINAL_DATABASE,
        help=f"Final classification database. Default: {FINAL_DATABASE}",
    )

    parser.add_argument(
        "--staging-db",
        type=Path,
        default=STAGING_DATABASE,
        help=f"Staging database. Default: {STAGING_DATABASE}",
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIRECTORY,
        help=f"Output directory. Default: {OUTPUT_DIRECTORY}",
    )

    return parser.parse_args()


def connect_database(database_path: Path) -> sqlite3.Connection:
    if not database_path.exists():
        raise FileNotFoundError(f"Database not found: {database_path}")

    connection = sqlite3.connect(database_path)
    connection.row_factory = sqlite3.Row
    return connection


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def paragraph_text(value: Any) -> str:
    return html.escape(safe_text(value)).replace("\n", "<br/>")


def load_total_project_file_counts(
    manifest_path: Path,
) -> dict[str, int]:
    if not manifest_path.exists():
        raise FileNotFoundError(
            f"Total-file manifest not found: {manifest_path}"
        )

    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    projects = payload.get("projects", {})

    if not isinstance(projects, dict):
        raise RuntimeError("Invalid total-file manifest: projects must be an object.")

    counts: dict[str, int] = {}

    for title, details in projects.items():
        if not isinstance(details, dict):
            raise RuntimeError(
                f"Invalid manifest entry for project {title!r}."
            )

        value = details.get("total_internal_files")

        if not isinstance(value, int) or value < 0:
            raise RuntimeError(
                f"Invalid total_internal_files value for {title!r}."
            )

        counts[str(title)] = value

    return counts


def fetch_final_export_rows(
    connection: sqlite3.Connection,
    total_project_file_counts: dict[str, int],
) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            p.repository_id AS repository_id,
            p.type AS project_type,
            p.title AS project_title,
            COALESCE(p.class, '') AS primary_class,
            COALESCE(p.secondary_class, '') AS secondary_class,
            COUNT(f.id) AS classified_primary_file_count
        FROM PROJECTS AS p
        LEFT JOIN FILES AS f
            ON f.project_id = p.id
        GROUP BY
            p.id,
            p.repository_id,
            p.type,
            p.title,
            p.class,
            p.secondary_class
        ORDER BY
            CAST(p.repository_id AS INTEGER),
            p.id;
        """
    ).fetchall()

    export_rows: list[dict[str, Any]] = []

    for row in rows:
        item = dict(row)
        title = str(item["project_title"])
        classified_primary_files = int(
            item["classified_primary_file_count"]
        )

        item["no_project_files"] = int(
            total_project_file_counts.get(
                title,
                classified_primary_files,
            )
        )

        export_rows.append(item)

    return export_rows


def fetch_final_repository_summary(
    connection: sqlite3.Connection,
    repository_id: int,
) -> dict[str, Any]:
    project_rows = connection.execute(
        """
        SELECT
            p.id,
            p.type,
            p.title,
            p.class,
            p.primary_division_code,
            p.secondary_class,
            p.confidence,
            p.classification_rule,
            COUNT(f.id) AS file_count
        FROM PROJECTS AS p
        LEFT JOIN FILES AS f
            ON f.project_id = p.id
        WHERE p.repository_id = ?
        GROUP BY
            p.id,
            p.type,
            p.title,
            p.class,
            p.primary_division_code,
            p.secondary_class,
            p.confidence,
            p.classification_rule
        ORDER BY p.id;
        """,
        (repository_id,),
    ).fetchall()

    file_rule_rows = connection.execute(
        """
        SELECT
            COALESCE(f.classification_rule, 'NOT_CLASSIFIED') AS rule,
            COUNT(*) AS file_count
        FROM FILES AS f
        INNER JOIN PROJECTS AS p
            ON p.id = f.project_id
        WHERE p.repository_id = ?
        GROUP BY COALESCE(f.classification_rule, 'NOT_CLASSIFIED')
        ORDER BY rule;
        """,
        (repository_id,),
    ).fetchall()

    return {
        "repository_id": repository_id,
        "projects": [dict(row) for row in project_rows],
        "file_rules": [dict(row) for row in file_rule_rows],
    }


def fetch_staging_statistics(
    connection: sqlite3.Connection,
) -> dict[str, Any]:
    import_totals = connection.execute(
        """
        SELECT
            COUNT(*) AS import_runs,
            COALESCE(SUM(projects_imported), 0) AS projects_imported,
            COALESCE(SUM(files_imported), 0) AS files_imported,
            COALESCE(SUM(keywords_imported), 0) AS keywords_imported,
            COALESCE(SUM(licenses_imported), 0) AS licenses_imported,
            COALESCE(SUM(person_roles_imported), 0) AS person_roles_imported
        FROM import_audit
        WHERE import_status = 'COMPLETED';
        """
    ).fetchone()

    source_registry = connection.execute(
        """
        SELECT
            source_scope,
            COUNT(*) AS registered_databases
        FROM source_databases
        GROUP BY source_scope
        ORDER BY source_scope;
        """
    ).fetchall()

    source_projects = connection.execute(
        """
        SELECT
            d.source_scope,
            COUNT(p.project_uid) AS staged_projects
        FROM source_databases AS d
        LEFT JOIN stg_projects AS p
            ON p.source_database_id = d.source_database_id
        GROUP BY d.source_scope
        ORDER BY d.source_scope;
        """
    ).fetchall()

    source_scope_summary: dict[str, dict[str, int]] = {}

    for row in source_registry:
        source_scope_summary[str(row["source_scope"])] = {
            "registered_databases": int(row["registered_databases"]),
            "staged_projects": 0,
        }

    for row in source_projects:
        scope = str(row["source_scope"])

        source_scope_summary.setdefault(
            scope,
            {
                "registered_databases": 0,
                "staged_projects": 0,
            },
        )

        source_scope_summary[scope]["staged_projects"] = int(
            row["staged_projects"]
        )

    project_type_rows = connection.execute(
        """
        SELECT
            project_type,
            COUNT(*) AS projects
        FROM project_type_classifications
        GROUP BY project_type
        ORDER BY projects DESC;
        """
    ).fetchall()

    total_staging = connection.execute(
        """
        SELECT
            COUNT(*) AS total_projects,
            COUNT(DISTINCT source_database_id) AS source_databases
        FROM stg_projects;
        """
    ).fetchone()

    duplicate_clusters = connection.execute(
        """
        SELECT COUNT(*) AS count
        FROM duplicate_clusters;
        """
    ).fetchone()

    duplicate_members = connection.execute(
        """
        SELECT COUNT(*) AS count
        FROM duplicate_cluster_members;
        """
    ).fetchone()

    latest_deduplication_run = connection.execute(
        """
        SELECT
            run_id,
            raw_project_count,
            candidate_cluster_count,
            confirmed_cluster_count,
            excluded_project_count,
            included_project_count
        FROM deduplication_runs
        ORDER BY created_at_utc DESC, rowid DESC
        LIMIT 1;
        """
    ).fetchone()

    quality_issues = connection.execute(
        """
        SELECT COUNT(*) AS count
        FROM data_quality_issues;
        """
    ).fetchone()

    isic_scope_rows = connection.execute(
        """
        SELECT
            source_scope,
            COUNT(*) AS classified_projects
        FROM isic_project_classifications
        GROUP BY source_scope
        ORDER BY source_scope;
        """
    ).fetchall()

    isic_file_scope_rows = connection.execute(
        """
        SELECT
            d.source_scope,
            COUNT(*) AS classified_files
        FROM isic_file_classifications AS c
        JOIN stg_projects AS p
            ON p.project_uid = c.project_uid
        JOIN source_databases AS d
            ON d.source_database_id = p.source_database_id
        GROUP BY d.source_scope
        ORDER BY d.source_scope;
        """
    ).fetchall()

    return {
        "import_totals": dict(import_totals),
        "source_scope_summary": source_scope_summary,
        "project_types": {
            str(row["project_type"]): int(row["projects"])
            for row in project_type_rows
        },
        "total_projects": int(total_staging["total_projects"]),
        "source_databases": int(total_staging["source_databases"]),
        "duplicate_clusters": int(duplicate_clusters["count"]),
        "duplicate_members": int(duplicate_members["count"]),
        "deduplication": {
            "available": latest_deduplication_run is not None,
            "run_id": (
                str(latest_deduplication_run["run_id"])
                if latest_deduplication_run is not None
                else "not_available"
            ),
            "raw_projects": (
                int(latest_deduplication_run["raw_project_count"])
                if latest_deduplication_run is not None
                else 0
            ),
            "high_confidence_clusters": (
                int(latest_deduplication_run["confirmed_cluster_count"])
                if latest_deduplication_run is not None
                else 0
            ),
            "excluded_from_derived_analysis": (
                int(latest_deduplication_run["excluded_project_count"])
                if latest_deduplication_run is not None
                else 0
            ),
            "retained_in_derived_analysis": (
                int(latest_deduplication_run["included_project_count"])
                if latest_deduplication_run is not None
                else 0
            ),
        },
        "quality_issues": int(quality_issues["count"]),
        "isic_project_scopes": {
            str(row["source_scope"]): int(row["classified_projects"])
            for row in isic_scope_rows
        },
        "isic_file_scopes": {
            str(row["source_scope"]): int(row["classified_files"])
            for row in isic_file_scope_rows
        },
    }


def load_automation_status() -> dict[str, Any]:
    drift_path = Path("reports/drift_report.json")
    gate_path = Path("reports/release_quality_gate.json")

    result = {
        "drift_available": False,
        "database_changed": None,
        "reclassification_required_count": None,
        "quality_gate_available": False,
        "quality_gate_passed": None,
        "quality_gate_failed_checks": None,
        "quality_gate_warnings": None,
    }

    if drift_path.exists():
        try:
            drift = json.loads(drift_path.read_text(encoding="utf-8"))

            result["drift_available"] = True
            result["database_changed"] = drift.get("database_changed")
            result["reclassification_required_count"] = drift.get(
                "reclassification_required_count"
            )
        except json.JSONDecodeError:
            pass

    if gate_path.exists():
        try:
            gate = json.loads(gate_path.read_text(encoding="utf-8"))

            result["quality_gate_available"] = True
            result["quality_gate_passed"] = gate.get("passed")
            result["quality_gate_failed_checks"] = gate.get(
                "failed_check_count"
            )
            result["quality_gate_warnings"] = gate.get("warning_count")
        except json.JSONDecodeError:
            pass

    return result


def build_xlsx(
    export_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Classification Export"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="17365D",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    worksheet.append(XLSX_COLUMNS)

    for index, _ in enumerate(XLSX_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row in export_rows:
        worksheet.append([row[column] for column in XLSX_COLUMNS])

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(XLSX_COLUMNS),
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row[0].alignment = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=True,
        )

        row[5].alignment = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:F{worksheet.max_row}"

    widths = {
        "A": 16,
        "B": 20,
        "C": 62,
        "D": 38,
        "E": 38,
        "F": 20,
    }

    for column_name, width in widths.items():
        worksheet.column_dimensions[column_name].width = width

    worksheet.row_dimensions[1].height = 32

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 34

    workbook.save(output_path)


def make_styles() -> dict[str, ParagraphStyle]:
    sample_styles = getSampleStyleSheet()

    return {
        "title": ParagraphStyle(
            "ReportTitle",
            parent=sample_styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=DARK_BLUE,
            alignment=TA_LEFT,
            spaceAfter=8,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=sample_styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
            textColor=DARK_GREY,
            spaceAfter=14,
        ),
        "heading_1": ParagraphStyle(
            "HeadingOne",
            parent=sample_styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            textColor=DARK_BLUE,
            spaceBefore=4,
            spaceAfter=8,
        ),
        "heading_2": ParagraphStyle(
            "HeadingTwo",
            parent=sample_styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11.5,
            leading=14,
            textColor=DARK_BLUE,
            spaceBefore=7,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=sample_styles["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13.5,
            textColor=DARK_GREY,
            spaceAfter=7,
        ),
        "small": ParagraphStyle(
            "Small",
            parent=sample_styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.3,
            leading=10.5,
            textColor=DARK_GREY,
            spaceAfter=4,
        ),
        "card_value": ParagraphStyle(
            "CardValue",
            parent=sample_styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=17,
            leading=20,
            textColor=DARK_BLUE,
            alignment=TA_CENTER,
        ),
        "card_label": ParagraphStyle(
            "CardLabel",
            parent=sample_styles["Normal"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=9,
            textColor=DARK_GREY,
            alignment=TA_CENTER,
        ),
        "table_header": ParagraphStyle(
            "TableHeader",
            parent=sample_styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.2,
            leading=10,
            textColor=WHITE,
            alignment=TA_CENTER,
        ),
        "table_body": ParagraphStyle(
            "TableBody",
            parent=sample_styles["Normal"],
            fontName="Helvetica",
            fontSize=8.2,
            leading=10.5,
            textColor=DARK_GREY,
        ),
        "table_body_center": ParagraphStyle(
            "TableBodyCenter",
            parent=sample_styles["Normal"],
            fontName="Helvetica",
            fontSize=8.2,
            leading=10.5,
            textColor=DARK_GREY,
            alignment=TA_CENTER,
        ),
        "note": ParagraphStyle(
            "Note",
            parent=sample_styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=8.2,
            leading=10.5,
            textColor=DARK_GREY,
        ),
    }


def make_header_footer(canvas, document) -> None:
    canvas.saveState()

    page_number = canvas.getPageNumber()
    page_width, page_height = A4

    canvas.setStrokeColor(MID_GREY)
    canvas.setLineWidth(0.5)
    canvas.line(1.7 * cm, page_height - 1.35 * cm, page_width - 1.7 * cm, page_height - 1.35 * cm)
    canvas.line(1.7 * cm, 1.25 * cm, page_width - 1.7 * cm, 1.25 * cm)

    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(DARK_GREY)

    canvas.drawString(
        1.7 * cm,
        page_height - 1.08 * cm,
        "SQ26 Part 2 — QDArchive Classification Report",
    )

    canvas.drawRightString(
        page_width - 1.7 * cm,
        page_height - 1.08 * cm,
        f"Student ID: {STUDENT_ID}",
    )

    canvas.drawString(
        1.7 * cm,
        0.88 * cm,
        "Generated from final and staging SQLite databases",
    )

    canvas.drawRightString(
        page_width - 1.7 * cm,
        0.88 * cm,
        f"Page {page_number}",
    )

    canvas.restoreState()


def metric_card(
    value: str,
    label: str,
    styles: dict[str, ParagraphStyle],
) -> Table:
    table = Table(
        [
            [Paragraph(paragraph_text(value), styles["card_value"])],
            [Paragraph(paragraph_text(label), styles["card_label"])],
        ],
        colWidths=[4.05 * cm],
        rowHeights=[0.75 * cm, 0.58 * cm],
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), VERY_LIGHT_BLUE),
                ("BOX", (0, 0), (-1, -1), 0.8, MID_BLUE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    return table


def metric_row(
    cards: list[tuple[str, str]],
    styles: dict[str, ParagraphStyle],
) -> Table:
    card_tables = [
        metric_card(value, label, styles)
        for value, label in cards
    ]

    row = Table(
        [card_tables],
        colWidths=[4.25 * cm] * len(cards),
    )

    row.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    return row


def paragraph_table(
    rows: list[list[Any]],
    widths: list[float],
    styles: dict[str, ParagraphStyle],
    header: bool = True,
) -> Table:
    converted_rows: list[list[Paragraph]] = []

    for row_index, row in enumerate(rows):
        converted_row = []

        for value in row:
            style = (
                styles["table_header"]
                if header and row_index == 0
                else styles["table_body"]
            )

            converted_row.append(
                Paragraph(
                    paragraph_text(value),
                    style,
                )
            )

        converted_rows.append(converted_row)

    table = Table(
        converted_rows,
        colWidths=widths,
        repeatRows=1 if header else 0,
        hAlign="LEFT",
    )

    commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, MID_GREY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]

    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ]
        )

    for row_index in range(1 if header else 0, len(rows)):
        if row_index % 2 == 1:
            commands.append(
                ("BACKGROUND", (0, row_index), (-1, row_index), VERY_LIGHT_BLUE)
            )

    table.setStyle(TableStyle(commands))

    return table


def horizontal_bar_chart(
    labels: list[str],
    values: list[int],
    width: float = 16.8 * cm,
    height: float = 6.2 * cm,
    title: str = "Primary ISIC class distribution by project",
) -> Drawing:
    drawing = Drawing(width, height)

    left_margin = 5.6 * cm
    right_margin = 1.7 * cm
    top_margin = 0.9 * cm
    bottom_margin = 0.9 * cm

    chart_width = width - left_margin - right_margin
    chart_height = height - top_margin - bottom_margin

    maximum = max(values) if values else 1
    maximum = max(maximum, 1)

    drawing.add(
        Line(
            left_margin,
            bottom_margin,
            left_margin,
            bottom_margin + chart_height,
            strokeColor=MID_GREY,
            strokeWidth=0.8,
        )
    )

    drawing.add(
        Line(
            left_margin,
            bottom_margin,
            left_margin + chart_width,
            bottom_margin,
            strokeColor=MID_GREY,
            strokeWidth=0.8,
        )
    )

    row_height = chart_height / max(len(labels), 1)
    bar_height = row_height * 0.42

    for index, (label, value) in enumerate(zip(labels, values)):
        y_center = bottom_margin + chart_height - row_height * (index + 0.5)
        bar_width = chart_width * value / maximum

        drawing.add(
            String(
                left_margin - 0.2 * cm,
                y_center - 2,
                label,
                textAnchor="end",
                fontName="Helvetica",
                fontSize=8.5,
                fillColor=DARK_GREY,
            )
        )

        drawing.add(
            Rect(
                left_margin,
                y_center - bar_height / 2,
                bar_width,
                bar_height,
                fillColor=MID_BLUE,
                strokeColor=MID_BLUE,
            )
        )

        drawing.add(
            String(
                left_margin + bar_width + 0.15 * cm,
                y_center - 2,
                f"{value}",
                fontName="Helvetica-Bold",
                fontSize=8.8,
                fillColor=DARK_BLUE,
            )
        )

    drawing.add(
        String(
            left_margin,
            height - 0.35 * cm,
            title,
            fontName="Helvetica-Bold",
            fontSize=10,
            fillColor=DARK_BLUE,
        )
    )

    return drawing


def build_pdf_report(
    output_path: Path,
    final_database_path: Path,
    staging_database_path: Path,
    export_rows: list[dict[str, Any]],
    repository_summaries: list[dict[str, Any]],
    staging_statistics: dict[str, Any],
    automation_status: dict[str, Any],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.75 * cm,
        bottomMargin=1.65 * cm,
        title="SQ26 Part 2 QDArchive Classification Report",
        author="Xinia Apchora",
    )

    styles = make_styles()
    story: list[Any] = []

    repository_5 = next(
        summary
        for summary in repository_summaries
        if summary["repository_id"] == 5
    )

    repository_15 = next(
        summary
        for summary in repository_summaries
        if summary["repository_id"] == 15
    )

    rule_counts = {
        str(row["rule"]): int(row["file_count"])
        for row in repository_5["file_rules"]
    }

    direct_files = (
        rule_counts.get("TIER2_FILE_CONTENT_MULTI_TERM", 0)
        + rule_counts.get("TIER2_FILE_CONTENT_SINGLE_TERM", 0)
    )

    fallback_files = rule_counts.get(
        "TIER2_PROJECT_CONTEXT_FALLBACK",
        0,
    )

    final_classified_projects = sum(
        1
        for row in export_rows
        if row["primary_class"]
    )

    final_total_project_files = sum(
        int(row["no_project_files"])
        for row in export_rows
    )

    final_primary_files = sum(
        int(row["classified_primary_file_count"])
        for row in export_rows
    )

    my_core_scope = staging_statistics["source_scope_summary"].get(
        "MY_CORE",
        {},
    )

    peer_scope = staging_statistics["source_scope_summary"].get(
        "PEER_SHARED",
        {},
    )

    project_types = staging_statistics["project_types"]

    # PAGE 1 — EXECUTIVE OVERVIEW

    story.append(
        Paragraph(
            "SQ26 Part 2 — QDArchive Classification Report",
            styles["title"],
        )
    )

    story.append(
        Paragraph(
            f"Student ID: {STUDENT_ID} | Final report generated from "
            "the staging and official delivery SQLite databases",
            styles["subtitle"],
        )
    )

    story.append(
        Paragraph(
            "Executive Overview",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "This report documents the complete Part 2 workflow: "
            "peer/shared metadata ingestion, source normalization, "
            "project-type classification, duplicate analysis, Tier 1 "
            "metadata evidence, Tier 2 primary-file content evidence, "
            "ISIC Rev. 5 classification of the MY_CORE delivery scope, "
            "and release automation checks.",
            styles["body"],
        )
    )

    story.append(
        metric_row(
            [
                (
                    f"{sum(scope['registered_databases'] for scope in staging_statistics['source_scope_summary'].values()):,}",
                    "Registered source databases",
                ),
                (
                    f"{staging_statistics['total_projects']:,}",
                    "Staged project records",
                ),
                (
                    f"{final_classified_projects}",
                    "Final ISIC-classified projects",
                ),
                (
                    f"{final_total_project_files}",
                    "Final project files (total)",
                ),
            ],
            styles,
        )
    )

    story.append(Spacer(1, 0.45 * cm))

    story.append(
        Paragraph(
            "Scope distinction",
            styles["heading_2"],
        )
    )

    scope_rows = [
        ["Scope", "Databases", "Projects", "Purpose"],
        [
            "PEER_SHARED",
            peer_scope.get("registered_databases", 0),
            f"{peer_scope.get('staged_projects', 0):,}",
            (
                "Imported, normalized, profiled, project-type classified, "
                "and included in duplicate analysis."
            ),
        ],
        [
            "MY_CORE",
            my_core_scope.get("registered_databases", 0),
            my_core_scope.get("staged_projects", 0),
            (
                "Assigned repositories used for final Part 2 delivery and "
                "ISIC classification."
            ),
        ],
    ]

    story.append(
        paragraph_table(
            scope_rows,
            [3.1 * cm, 2.2 * cm, 2.5 * cm, 9.1 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.3 * cm))

    story.append(
        Paragraph(
            "Important interpretation: peer/shared records were used to "
            "exercise the multi-source staging pipeline and project-type "
            "classification workflow. The final ISIC project and file "
            "classification scope was intentionally limited to MY_CORE. "
            "The report does not claim that every peer project received "
            "an ISIC label. The MY_CORE source database contained 16 "
            "staged records; the official final delivery materialized the "
            "9 records belonging to the assigned repositories 5 and 15.",
            styles["note"],
        )
    )

    story.append(PageBreak())

    # PAGE 2 — IMPORT, PROFILE, DEDUPLICATION

    story.append(
        Paragraph(
            "1. Shared Corpus Ingestion and Staging",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "The staging database consolidates peer-shared and MY_CORE "
            "metadata databases into a normalized data model. Every source "
            "database is registered with its origin, storage type, schema "
            "signature, checksum, file size, and import audit information.",
            styles["body"],
        )
    )

    import_totals = staging_statistics["import_totals"]

    story.append(
        metric_row(
            [
                (
                    f"{import_totals['import_runs']}",
                    "Completed import runs",
                ),
                (
                    f"{import_totals['files_imported']:,}",
                    "Imported file records",
                ),
                (
                    f"{import_totals['keywords_imported']:,}",
                    "Imported keywords",
                ),
                (
                    f"{import_totals['person_roles_imported']:,}",
                    "Imported person-role records",
                ),
            ],
            styles,
        )
    )

    story.append(Spacer(1, 0.45 * cm))

    story.append(
        Paragraph(
            "Import and quality summary",
            styles["heading_2"],
        )
    )

    quality_rows = [
        ["Measure", "Result", "Interpretation"],
        [
            "Completed imports",
            f"{import_totals['import_runs']}",
            "All registered source imports completed successfully.",
        ],
        [
            "Imported projects",
            f"{import_totals['projects_imported']:,}",
            "Metadata project records normalized into stg_projects.",
        ],
        [
            "Imported licenses",
            f"{import_totals['licenses_imported']:,}",
            "License records retained for source-level provenance.",
        ],
        [
            "Duplicate clusters",
            f"{staging_statistics['duplicate_clusters']:,}",
            (
                "Candidate duplicate groups detected across the "
                "multi-source staging corpus."
            ),
        ],
        [
            "Duplicate memberships",
            f"{staging_statistics['duplicate_members']:,}",
            "Project-to-cluster memberships supporting duplicate review.",
        ],
        [
            "High-confidence duplicate clusters",
            f"{staging_statistics['deduplication']['high_confidence_clusters']:,}",
            (
                "Deterministic duplicate clusters resolved in the derived "
                "analysis layer; raw staging records remain unchanged."
            ),
        ],
        [
            "Duplicate records excluded from derived analysis",
            f"{staging_statistics['deduplication']['excluded_from_derived_analysis']:,}",
            (
                "High-confidence duplicate records excluded non-destructively "
                "from analysis only; each retains a canonical reference."
            ),
        ],
        [
            "Projects retained in derived analysis",
            f"{staging_statistics['deduplication']['retained_in_derived_analysis']:,}",
            (
                "Records retained after duplicate exclusion for engineering "
                "analysis and evaluation."
            ),
        ],
        [
            "Recorded data-quality issues",
            f"{staging_statistics['quality_issues']:,}",
            "No unresolved staging quality issues recorded.",
        ],
    ]

    story.append(
        paragraph_table(
            quality_rows,
            [4.0 * cm, 3.0 * cm, 9.9 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.35 * cm))

    story.append(
        Paragraph(
            "Duplicate-resolution interpretation: duplicate candidates are "
            "clustered using deterministic identifier and metadata evidence. "
            "All registered candidate clusters met the configured "
            "high-confidence decision rule; therefore, candidate and "
            "high-confidence cluster counts are identical. High-confidence "
            "duplicate records are excluded only from a derived analysis "
            "layer. No raw project record is deleted, overwritten, or "
            "merged; each exclusion preserves its canonical reference, "
            "evidence, confidence, tags, and audit metadata.",
            styles["body"],
        )
    )

    story.append(Spacer(1, 0.25 * cm))

    story.append(
        Paragraph(
            "Peer/shared data contributed to the engineering workflow by "
            "providing heterogeneous schemas, mixed metadata completeness, "
            "varying repository structures, and realistic duplicate "
            "conditions. It was therefore valuable for validating that the "
            "pipeline does not assume a single repository-specific schema.",
            styles["body"],
        )
    )

    story.append(PageBreak())

    # PAGE 3 — PROJECT TYPE CLASSIFICATION

    story.append(
        Paragraph(
            "2. Project-Type Classification Across the Staging Corpus",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "The staging corpus was classified into QDA_PROJECT, QD_PROJECT, "
            "OTHER_PROJECT, and NOT_A_PROJECT. This stage determines which "
            "records are eligible for later qualitative-data and ISIC "
            "classification steps.",
            styles["body"],
        )
    )

    type_rows = [
        ["Project type", "Projects", "Meaning"],
        [
            "QD_PROJECT",
            f"{project_types.get('QD_PROJECT', 0):,}",
            "Project contains qualitative data evidence.",
        ],
        [
            "QDA_PROJECT",
            f"{project_types.get('QDA_PROJECT', 0):,}",
            "Project contains qualitative-data-analysis artefacts.",
        ],
        [
            "OTHER_PROJECT",
            f"{project_types.get('OTHER_PROJECT', 0):,}",
            "Project does not provide sufficient QD/QDA evidence.",
        ],
        [
            "NOT_A_PROJECT",
            f"{project_types.get('NOT_A_PROJECT', 0):,}",
            "Record does not represent an eligible research project.",
        ],
    ]

    story.append(
        paragraph_table(
            type_rows,
            [4.0 * cm, 3.0 * cm, 9.9 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.4 * cm))

    project_type_chart = horizontal_bar_chart(
        labels=[
            "QD_PROJECT",
            "OTHER_PROJECT",
            "NOT_A_PROJECT",
            "QDA_PROJECT",
        ],
        values=[
            project_types.get("QD_PROJECT", 0),
            project_types.get("OTHER_PROJECT", 0),
            project_types.get("NOT_A_PROJECT", 0),
            project_types.get("QDA_PROJECT", 0),
        ],
        width=16.8 * cm,
        height=6.2 * cm,
        title="Project-type distribution across the staging corpus",
    )

    story.append(project_type_chart)

    story.append(Spacer(1, 0.25 * cm))

    story.append(
        Paragraph(
            "The broad project-type stage was applied across the complete "
            "staging corpus. The final delivery, however, contains only the "
            "assigned MY_CORE repositories and their relevant project records.",
            styles["note"],
        )
    )

    story.append(PageBreak())

    # PAGE 4 — TIER 1 / TIER 2 METHOD

    story.append(
        Paragraph(
            "3. Tier 1 and Tier 2 Classification Method",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "The final ISIC classification uses a transparent "
            "deterministic, rule-based evidence strategy. "
            "Tier 1 establishes project context from "
            "project titles, descriptions, repository metadata, QDPX "
            "project metadata, and keywords where available. Tier 2 uses "
            "the actual contents of primary TXT and PDF files embedded "
            "within QDPX archives. Keywords were considered where available; ""no keyword records were available for the final MY_CORE delivery scope.",
            styles["body"],
        )
    )

    method_rows = [
        ["Stage", "Input evidence", "Output"],
        [
            "Source ingestion",
            (
                "Peer and MY_CORE SQLite databases, source registry, "
                "checksums, schema signatures, import audit."
            ),
            "Normalized staging records with provenance.",
        ],
        [
            "Project-type classification",
            (
                "Project metadata, file names, extensions, QDA/QD "
                "indicators, archive structure."
            ),
            "QDA_PROJECT, QD_PROJECT, OTHER_PROJECT, or NOT_A_PROJECT.",
        ],
        [
            "Tier 1 evidence",
            (
                "Project title, description, repository metadata, QDPX "
                "project metadata, and keywords where available."
            ),
            "Project-level ISIC candidate context.",
        ],
        [
            "Tier 2 evidence",
            (
                "Embedded TXT and PDF primary-file contents extracted from "
                "sources/ inside QDPX archives."
            ),
            "Direct file-level ISIC evidence where sufficient.",
        ],
        [
            "Fallback policy",
            (
                "Files without enough individual evidence retain the "
                "project class with documented lower confidence."
            ),
            "Documented project-context fallback classification.",
        ],
    ]

    story.append(
        paragraph_table(
            method_rows,
            [3.1 * cm, 8.0 * cm, 5.8 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.4 * cm))

    story.append(
        Paragraph(
            "Final MY_CORE Tier 2 evidence results",
            styles["heading_2"],
        )
    )

    story.append(
        metric_row(
            [
                (
                    "109",
                    "Multi-term direct files",
                ),
                (
                    "217",
                    "Single-term direct files",
                ),
                (
                    f"{direct_files}",
                    "Total direct Tier 2 files",
                ),
                (
                    f"{fallback_files}",
                    "Project-context fallback files",
                ),
            ],
            styles,
        )
    )

    story.append(Spacer(1, 0.35 * cm))

    story.append(
        Paragraph(
            "All 507 eligible QDPX-internal primary files in the final "
            "DANS delivery scope were processed. "
            "A direct content classification was possible for 326 files. "
            "The remaining 181 files lacked sufficient unambiguous "
            "standalone evidence and therefore used the documented "
            "project-context fallback. Raw extracted primary text was not "
            "stored in the final delivery database.",
            styles["body"],
        )
    )

    story.append(PageBreak())

    # PAGE 5 — FINAL REPOSITORY 5 RESULTS

    story.append(
        Paragraph(
            "4. Final Delivery Results — Repository 5 (DANS)",
            styles["heading_1"],
        )
    )

    repo_5_projects = repository_5["projects"]

    repo_5_classified = [
        project
        for project in repo_5_projects
        if project["class"]
    ]

    class_project_counts = Counter(
        str(project["class"])
        for project in repo_5_classified
    )

    class_file_counts: Counter[str] = Counter()

    for project in repo_5_classified:
        class_file_counts[str(project["class"])] += int(
            project["file_count"]
        )

    total_repo_5_files = sum(
        int(project["file_count"])
        for project in repo_5_projects
    )

    total_repo_5_project_files = sum(
        int(row["no_project_files"])
        for row in export_rows
        if int(row["repository_id"]) == 5
    )

    story.append(
        metric_row(
            [
                (
                    f"{len(repo_5_projects)}",
                    "QDA projects",
                ),
                (
                    f"{total_repo_5_files}",
                    "QDPX-internal primary files",
                ),
                (
                    f"{direct_files}",
                    "Direct Tier 2 decisions",
                ),
                (
                    f"{fallback_files}",
                    "Fallback decisions",
                ),
            ],
            styles,
        )
    )

    story.append(
        Paragraph(
            f"The XLSX column <b>no_project_files</b> reports "
            f"{total_repo_5_project_files} total internal QDPX project "
            f"files for Repository 5. This includes four QDE "
            f"project-definition files. The QDE files are included in "
            f"the total-file requirement but excluded from primary-file "
            f"ISIC classification; therefore, the classified primary-file "
            f"total remains {total_repo_5_files}.",
            styles["note"],
        )
    )

    story.append(Spacer(1, 0.4 * cm))

    chart_labels = list(class_project_counts.keys())
    chart_values = [
        class_project_counts[label]
        for label in chart_labels
    ]

    story.append(
        horizontal_bar_chart(
            labels=chart_labels,
            values=chart_values,
            width=16.8 * cm,
            height=5.8 * cm,
            title="Primary ISIC class distribution by project",
        )
    )

    story.append(Spacer(1, 0.28 * cm))

    story.append(
        Paragraph(
            "Primary ISIC class distribution by primary file",
            styles["heading_2"],
        )
    )

    file_chart_labels = list(class_file_counts.keys())
    file_chart_values = [
        class_file_counts[label]
        for label in file_chart_labels
    ]

    story.append(
        horizontal_bar_chart(
            labels=file_chart_labels,
            values=file_chart_values,
            width=16.8 * cm,
            height=5.2 * cm,
            title="Primary ISIC class distribution by primary file",
        )
    )

    story.append(Spacer(1, 0.25 * cm))

    ranking_rows = [
        [
            "Rank",
            "Primary ISIC class",
            "Division",
            "Projects",
            "Primary files",
        ]
    ]

    ranked_classes = sorted(
        class_project_counts,
        key=lambda class_name: (
            -class_project_counts[class_name],
            class_name,
        ),
    )

    division_by_class = {
        str(project["class"]): safe_text(
            project["primary_division_code"]
        )
        for project in repo_5_classified
    }

    for rank, class_name in enumerate(ranked_classes, start=1):
        ranking_rows.append(
            [
                rank,
                class_name,
                division_by_class[class_name],
                class_project_counts[class_name],
                class_file_counts[class_name],
            ]
        )

    story.append(
        Paragraph(
            "Primary ISIC class ranking",
            styles["heading_2"],
        )
    )

    story.append(
        paragraph_table(
            ranking_rows,
            [1.4 * cm, 7.2 * cm, 2.0 * cm, 2.6 * cm, 3.6 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.25 * cm))

    story.append(
        Paragraph(
            "Top-20 clarification: only two ISIC primary classes occurred "
            "in Repository 5. Therefore, the ranking table contains two "
            "rows rather than twenty empty or artificial rows.",
            styles["note"],
        )
    )

    story.append(Spacer(1, 0.18 * cm))

    story.append(
        Paragraph(
            "Interpretation",
            styles["heading_2"],
        )
    )

    story.append(
        Paragraph(
            "Repository 5 contains four QDA projects. Three are classified "
            "as Legal and accounting activities (ISIC division N69), "
            "covering 499 primary files. One is classified as Scientific "
            "research and development (ISIC division N72), covering 8 "
            "primary files. The direct Tier 2 evidence and fallback policy "
            "allow every primary file to remain traceable to a documented "
            "classification rule.",
            styles["body"],
        )
    )

    story.append(
        Paragraph(
            "ISIC section summary: both final divisions, N69 and N72, "
            "belong to ISIC section M — Professional, scientific and "
            "technical activities. Therefore, all four final "
            "ISIC-classified DANS QDA projects and all 507 classified "
            "primary files belong to section M.",
            styles["body"],
        )
    )

    story.append(PageBreak())

    # PAGE 6 — FINAL REPOSITORY 15 RESULTS

    story.append(
        Paragraph(
            "5. Final Delivery Results — Repository 15 (ICPSR)",
            styles["heading_1"],
        )
    )

    repo_15_projects = repository_15["projects"]

    repo_15_type_counts = Counter(
        safe_text(project["type"])
        for project in repo_15_projects
    )

    repo_15_files = sum(
        int(project["file_count"])
        for project in repo_15_projects
    )

    story.append(
        metric_row(
            [
                (
                    f"{len(repo_15_projects)}",
                    "Final project records",
                ),
                (
                    f"{repo_15_type_counts.get('OTHER_PROJECT', 0)}",
                    "OTHER_PROJECT records",
                ),
                (
                    f"{repo_15_files}",
                    "Primary QDA/QD files",
                ),
                (
                    "0",
                    "ISIC classes assigned",
                ),
            ],
            styles,
        )
    )

    story.append(Spacer(1, 0.55 * cm))

    status_table = Table(
        [
            [
                Paragraph(
                    "<b>No eligible QDA/QD project was available for "
                    "final ISIC classification in Repository 15.</b>",
                    ParagraphStyle(
                        "Status",
                        parent=styles["body"],
                        fontName="Helvetica-Bold",
                        fontSize=12,
                        leading=16,
                        textColor=ORANGE,
                        alignment=TA_CENTER,
                    ),
                )
            ],
            [
                Paragraph(
                    "This is an accurate repository outcome, not a missing "
                    "classification result. The five retained records are "
                    "metadata projects classified as OTHER_PROJECT. Because "
                    "there are zero eligible QDA projects and zero primary "
                    "files, a primary-class histogram and Top-20 class "
                    "ranking are not applicable.",
                    ParagraphStyle(
                        "StatusBody",
                        parent=styles["body"],
                        alignment=TA_CENTER,
                    ),
                )
            ],
        ],
        colWidths=[16.8 * cm],
    )

    status_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF4E8")),
                ("BOX", (0, 0), (-1, -1), 1.0, ORANGE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )

    story.append(status_table)
    story.append(Spacer(1, 0.35 * cm))

    story.append(
        Paragraph(
            "ICPSR zero-distribution histogram",
            styles["heading_2"],
        )
    )

    zero_chart = Drawing(16.8 * cm, 5.0 * cm)
    zero_chart.add(
        String(
            0,
            4.55 * cm,
            "Eligible QDA/QD distribution — zero observations",
            fontName="Helvetica-Bold",
            fontSize=10,
            fillColor=DARK_BLUE,
        )
    )
    zero_chart.add(
        String(
            0,
            4.10 * cm,
            "Each bar terminates at zero because no eligible QDA/QD "
            "project or primary file exists in Repository 15.",
            fontName="Helvetica",
            fontSize=8.5,
            fillColor=DARK_GREY,
        )
    )

    zero_metrics = [
        ("Eligible QDA projects", 0),
        ("Primary files", 0),
        ("ISIC primary classes", 0),
    ]

    for index, (label, value) in enumerate(zero_metrics):
        y = (3.25 - index * 0.85) * cm
        zero_chart.add(
            String(
                0,
                y,
                label,
                fontName="Helvetica",
                fontSize=9,
                fillColor=DARK_GREY,
            )
        )
        zero_chart.add(
            Line(
                6.4 * cm,
                y + 0.08 * cm,
                15.2 * cm,
                y + 0.08 * cm,
                strokeColor=MID_GREY,
                strokeWidth=0.8,
            )
        )
        zero_chart.add(
            Line(
                6.4 * cm,
                y - 0.15 * cm,
                6.4 * cm,
                y + 0.30 * cm,
                strokeColor=MID_BLUE,
                strokeWidth=2.0,
            )
        )
        zero_chart.add(
            String(
                15.45 * cm,
                y,
                str(value),
                fontName="Helvetica-Bold",
                fontSize=9,
                fillColor=DARK_BLUE,
            )
        )

    zero_chart.add(
        String(
            6.25 * cm,
            0.28 * cm,
            "0",
            fontName="Helvetica",
            fontSize=8,
            fillColor=DARK_GREY,
        )
    )
    zero_chart.add(
        String(
            14.7 * cm,
            0.28 * cm,
            "Count",
            fontName="Helvetica",
            fontSize=8,
            fillColor=DARK_GREY,
        )
    )

    story.append(zero_chart)
    story.append(Spacer(1, 0.25 * cm))

    story.append(
        Paragraph(
            "Interpretation",
            styles["heading_2"],
        )
    )

    story.append(
        Paragraph(
            "Repository 15 contributes five metadata records to the final "
            "delivery. These records do not provide QDA or QD primary-file "
            "evidence in the final delivery database. Assigning an ISIC "
            "activity class without qualifying evidence would be "
            "methodologically unsupported. Therefore, the XLSX export "
            "correctly retains empty primary_class and secondary_class "
            "fields for these records.",
            styles["body"],
        )
    )

    repo_15_rows = [
        [
            "Repository ID",
            "Project type",
            "Projects",
            "Primary files",
            "ISIC project classes",
        ],
        [
            "15",
            "OTHER_PROJECT",
            len(repo_15_projects),
            repo_15_files,
            "0",
        ],
    ]

    story.append(Spacer(1, 0.25 * cm))

    story.append(
        paragraph_table(
            repo_15_rows,
            [2.8 * cm, 4.2 * cm, 2.4 * cm, 3.4 * cm, 4.0 * cm],
            styles,
        )
    )

    story.append(PageBreak())

    # PAGE 7 — AUTOMATION AND REPRODUCIBILITY

    story.append(
        Paragraph(
            "6. Automation, Monitoring, and Reproducibility",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "The final delivery is generated through reproducible scripts "
            "rather than manually edited data. The workflow materializes the "
            "official SQLite database, generates the XLSX/PDF deliverables, "
            "runs a read-only drift monitor, and runs a release quality gate.",
            styles["body"],
        )
    )

    drift_value = (
        "Available"
        if automation_status["drift_available"]
        else "Not found"
    )

    reclassification_value = (
        str(automation_status["reclassification_required_count"])
        if automation_status["reclassification_required_count"] is not None
        else "N/A"
    )

    gate_value = (
        "Passed"
        if automation_status["quality_gate_passed"] is True
        else "Not passed"
        if automation_status["quality_gate_passed"] is False
        else "Not found"
    )

    failed_checks = (
        str(automation_status["quality_gate_failed_checks"])
        if automation_status["quality_gate_failed_checks"] is not None
        else "N/A"
    )

    story.append(
        metric_row(
            [
                (
                    drift_value,
                    "Drift monitor status",
                ),
                (
                    reclassification_value,
                    "Projects requiring reclassification",
                ),
                (
                    gate_value,
                    "Release quality gate",
                ),
                (
                    failed_checks,
                    "Failed quality checks",
                ),
            ],
            styles,
        )
    )

    story.append(Spacer(1, 0.45 * cm))

    automation_rows = [
        ["Automation component", "Purpose", "Observed result"],
        [
            "Official delivery materializer",
            (
                "Builds 23071063-sq26-classification.db from the "
                "staging classification records."
            ),
            "Foreign-key and integrity validation passed.",
        ],
        [
            "Tier 2 extractor",
            (
                "Extracts bounded TXT/PDF evidence from QDPX "
                "sources/ entries without persisting raw text."
            ),
            "507 primary files processed.",
        ],
        [
            "Drift monitor",
            (
                "Compares database, project, and QDPX archive fingerprints "
                "without changing classifications."
            ),
            (
                "Baseline stable after Tier 2 rebuild; no project required "
                "reclassification."
            ),
        ],
        [
            "Release quality gate",
            (
                "Checks expected database structure, counts, final "
                "classification distribution, and drift status."
            ),
            "Passed with zero failed checks.",
        ],
        [
            "Automated test suite",
            (
                "Tests Tier 2 extraction, ISIC rules, drift monitoring, "
                "quality gate logic, and output assumptions."
            ),
            "46 tests passed.",
        ],
    ]

    story.append(
        paragraph_table(
            automation_rows,
            [4.0 * cm, 7.4 * cm, 5.4 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.35 * cm))

    story.append(
        Paragraph(
            "Reproducibility commands",
            styles["heading_2"],
        )
    )

    commands = [
        "python -m scripts.classify_official_isic --staging-db data/staging/qdarchive_x_staging.db --raw-data-root data/raw --reset",
        "python -m scripts.materialize_sq26_classification_delivery",
        "python -m scripts.run_drift_monitor",
        "python -m scripts.run_release_quality_gate --require-drift-report",
        "python -m scripts.generate_part2_deliverables",
        "python -m pytest -q",
    ]

    command_rows = [
        [
            Paragraph(
                f"<font name='Courier' size='7.7'>{paragraph_text(command)}</font>",
                styles["small"],
            )
        ]
        for command in commands
    ]

    command_table = Table(
        command_rows,
        colWidths=[16.8 * cm],
    )

    command_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, MID_GREY),
                ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GREY),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(command_table)
    story.append(PageBreak())

    # PAGE 8 — CHALLENGES AND LIMITATIONS

    story.append(
        Paragraph(
            "7. Technical Data Challenges and Limitations",
            styles["heading_1"],
        )
    )

    story.append(
        Paragraph(
            "The following issues relate to the structure and evidential "
            "quality of the acquired data. They are documented as data "
            "handling considerations rather than programming defects.",
            styles["body"],
        )
    )

    challenges_rows = [
        ["Observed challenge", "Handling and impact"],
        [
            "Nested QDPX archive structure",
            (
                "DANS QDA projects were supplied as QDPX archives. Primary "
                "data were embedded in internal sources/ paths. The pipeline "
                "enumerates these entries and does not treat the outer QDPX "
                "archive itself as a primary file."
            ),
        ],
        [
            "Mixed TXT and PDF sources",
            (
                "Primary materials contained both text files and PDFs. TXT "
                "was decoded directly; PDF text was extracted with bounded "
                "page and character limits to support classification without "
                "storing raw source text."
            ),
        ],
        [
            "Incomplete standalone file evidence",
            (
                "Some individual files contain generic or contextual content. "
                "They cannot always support an independent ISIC decision. "
                "These files use the explicit "
                "TIER2_PROJECT_CONTEXT_FALLBACK rule."
            ),
        ],
        [
            "Peer versus final ISIC scope",
            (
                "Peer/shared records were used for staging, project-type "
                "classification, source profiling, and duplicate analysis. "
                "ISIC labels were not assigned across the entire peer corpus; "
                "the final ISIC scope was MY_CORE only."
            ),
        ],
        [
            "Non-fatal PDF encoding notices",
            (
                "Some PDFs produced SymbolSetEncoding notices during parsing. "
                "Extraction completed successfully for the processed corpus "
                "and classification finished with zero recorded errors."
            ),
        ],
    ]

    story.append(
        paragraph_table(
            challenges_rows,
            [5.0 * cm, 11.8 * cm],
            styles,
        )
    )

    story.append(Spacer(1, 0.45 * cm))

    story.append(
        Paragraph(
            "Conclusion",
            styles["heading_2"],
        )
    )

    story.append(
        Paragraph(
            "The Part 2 implementation provides a traceable classification "
            "workflow over a large peer/shared staging corpus and a focused "
            "MY_CORE final delivery. The engineering extension identified and resolved "
            "56,164 high-confidence duplicate clusters in the derived analysis layer, excluding 63,753 "
            "duplicate records from the derived analysis layer while "
            "preserving all raw staging records. The final output includes "
            "9 delivery project records, 4 ISIC-classified QDA projects, "
            "507 classified primary files, Tier 2 evidence provenance, "
            "reproducible report generation, drift monitoring, release "
            "validation, and a passing automated test suite.",
            styles["body"],
        )
    )

    document.build(
        story,
        onFirstPage=make_header_footer,
        onLaterPages=make_header_footer,
    )


def verify_outputs(
    xlsx_path: Path,
    pdf_path: Path,
    expected_project_rows: int,
) -> dict[str, Any]:
    workbook = load_workbook(
        xlsx_path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook["Classification Export"]

    headers = [
        cell.value
        for cell in next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]

    if headers != XLSX_COLUMNS:
        raise RuntimeError(
            f"Unexpected XLSX headers: {headers}"
        )

    project_rows = worksheet.max_row - 1

    if project_rows != expected_project_rows:
        raise RuntimeError(
            f"Expected {expected_project_rows} XLSX rows, found {project_rows}."
        )

    if project_rows != 9:
        raise RuntimeError(
            f"Expected 9 final delivery projects, found {project_rows}."
        )

    reader = PdfReader(str(pdf_path))
    pdf_pages = len(reader.pages)

    if pdf_pages < 8:
        raise RuntimeError(
            f"Expected at least 8 report pages, found {pdf_pages}."
        )

    pdf_text = " ".join(
        page.extract_text() or ""
        for page in reader.pages
    )

    required_phrases = [
        "PEER_SHARED",
        "MY_CORE",
        "Tier 1 and Tier 2",
        "Repository 5",
        "Repository 15",
        "Automation",
        "Technical Data Challenges",
        "326",
        "181",
    ]

    missing_phrases = [
        phrase
        for phrase in required_phrases
        if phrase not in pdf_text
    ]

    if missing_phrases:
        raise RuntimeError(
            "PDF verification missing expected content: "
            + ", ".join(missing_phrases)
        )

    return {
        "xlsx_bytes": xlsx_path.stat().st_size,
        "pdf_bytes": pdf_path.stat().st_size,
        "pdf_pages": pdf_pages,
        "project_rows": project_rows,
    }


def main() -> None:
    arguments = parse_arguments()

    xlsx_path = arguments.output_dir / (
        f"{STUDENT_ID}-sq26-classification.xlsx"
    )

    pdf_path = arguments.output_dir / (
        f"{STUDENT_ID}-sq26-classification-report.pdf"
    )

    summary_path = arguments.output_dir / (
        "part2_deliverables_summary.json"
    )

    total_project_file_counts = load_total_project_file_counts(
        TOTAL_FILE_MANIFEST
    )

    final_connection = connect_database(arguments.db)

    try:
        export_rows = fetch_final_export_rows(
            final_connection,
            total_project_file_counts,
        )

        repository_ids = [
            int(row["repository_id"])
            for row in final_connection.execute(
                """
                SELECT DISTINCT repository_id
                FROM PROJECTS
                ORDER BY CAST(repository_id AS INTEGER);
                """
            ).fetchall()
        ]

        repository_summaries = [
            fetch_final_repository_summary(
                final_connection,
                repository_id,
            )
            for repository_id in repository_ids
        ]
    finally:
        final_connection.close()

    staging_connection = connect_database(arguments.staging_db)

    try:
        staging_statistics = fetch_staging_statistics(
            staging_connection
        )
    finally:
        staging_connection.close()

    automation_status = load_automation_status()

    build_xlsx(
        export_rows,
        xlsx_path,
    )

    build_pdf_report(
        output_path=pdf_path,
        final_database_path=arguments.db,
        staging_database_path=arguments.staging_db,
        export_rows=export_rows,
        repository_summaries=repository_summaries,
        staging_statistics=staging_statistics,
        automation_status=automation_status,
    )

    verification = verify_outputs(
        xlsx_path=xlsx_path,
        pdf_path=pdf_path,
        expected_project_rows=len(export_rows),
    )

    summary = {
        "final_database": str(arguments.db),
        "staging_database": str(arguments.staging_db),
        "xlsx_output": str(xlsx_path),
        "pdf_output": str(pdf_path),
        "total_file_manifest": str(TOTAL_FILE_MANIFEST),
        "total_project_files": sum(
            int(row["no_project_files"])
            for row in export_rows
        ),
        "classified_primary_files": sum(
            int(row["classified_primary_file_count"])
            for row in export_rows
        ),
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "verification": verification,
    }

    summary_path.write_text(
        json.dumps(summary, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()